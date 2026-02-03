# created : 2025-10-23  컨테이너 실행
# modified : 2025-10-27 +-20% 알람 해제
# modified : 2025-10-27 로그파일 월단위 설정
# modified : 2025-10-27 메시지 형식 수정 (10%, 15%는 5%이상에 포함)
# modified : 2026-02-03 종목별 감시(upbitMA.list.xlsx) 추가
# modified : 2026-02-03 설정 전부 .env 사용

import requests
import time
import datetime
import os
import sys
import re
import atexit
import signal

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

from dotenv import load_dotenv

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(SCRIPT_DIR, ".env"))

# 설정값 불러오기 (.env)
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()
TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID", "").strip()
ALL_MA_INTERVAL = int(os.getenv("ALL_MA_INTERVAL", "3600").strip() or "3600")  # 전체 종목 분석 주기(초)
LIST_MA_INTERVAL = int(os.getenv("LIST_MA_INTERVAL", "60").strip() or "60")  # 종목별 감시 주기(초), 기본 1분
LIST_FILE_RAW = os.getenv("LIST_FILE", "").strip()
if LIST_FILE_RAW:
    EXCEL_LIST_PATH = os.path.join(SCRIPT_DIR, LIST_FILE_RAW) if not os.path.isabs(LIST_FILE_RAW) else LIST_FILE_RAW
else:
    EXCEL_LIST_PATH = None

if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
    raise ValueError("TELEGRAM_BOT_TOKEN, TELEGRAM_CHAT_ID가 .env에 필요합니다.")

# 종목별 감시용 캐시: 마켓 매핑 + 전종목 시세 (API 호출 최소화)
_MARKET_CACHE_TTL = 600  # 초 (10분)
_market_map_cache = None
_krw_markets_cache = None
_market_cache_time = 0

# 종목별 감시: 한 번 알림 보낸 (종목명, 감시사유)는 이후 감시 대상에서 제외 (감시중 X와 동일)
_list_alert_sent = set()

# ✅ 실행 시마다 날짜 확인 → 파일명 동적으로 갱신
TODAY = datetime.date.today().strftime("%Y%m%d")
TODAY_MONTH = datetime.date.today().strftime("%Y%m")
SCRIPT_FILENAME = os.path.splitext(os.path.basename(sys.argv[0]))[0]
# LOG_DIR_FILENAME = os.path.join(SCRIPT_DIR, f"{SCRIPT_FILENAME}_{TODAY}.md")
LOG_DIR_FILENAME = os.path.join(SCRIPT_DIR, f"{SCRIPT_FILENAME}_{TODAY_MONTH}.md")



def send_telegram_message(message):
    """텔레그램 알림 전송"""
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    payload = {"chat_id": TELEGRAM_CHAT_ID, "text": message}
    try:
        r = requests.post(url, data=payload, timeout=10)
        if r.status_code != 200:
            print(f"[텔레그램 전송 실패] HTTP {r.status_code}: {r.text[:200]}")
    except Exception as e:
        print(f"[텔레그램 전송 실패] {e}")

def get_upbit_markets():
    """업비트 원화시장 종목 목록 가져오기"""
    url = "https://api.upbit.com/v1/market/all"
    res = requests.get(url).json()
    return [m['market'] for m in res if m['market'].startswith('KRW-')]


def get_upbit_markets_all():
    """업비트 마켓 전체 조회 (종목명→마켓코드 매핑용)"""
    url = "https://api.upbit.com/v1/market/all"
    resp = requests.get(url, params={"isDetails": "true"}, timeout=10)
    resp.raise_for_status()
    return resp.json()


def build_name_market_map():
    """종목명/심볼 → 마켓코드(KRW-XXX) 매핑 생성"""
    markets = get_upbit_markets_all()
    name_map = {}
    for m in markets:
        mkt = m["market"]
        if not mkt.startswith("KRW-"):
            continue
        korean = m.get("korean_name", "")
        english = m.get("english_name", "")
        symbol = mkt.replace("KRW-", "")
        if korean:
            name_map[korean] = mkt
        if english:
            name_map[english] = mkt
        name_map[symbol] = mkt
        name_map[mkt] = mkt
        name_map[f"{symbol}/KRW"] = mkt
    return name_map


def get_cached_market_data():
    """종목명 매핑 + KRW 마켓 목록 캐시. TTL 내에는 API 호출 없이 반환."""
    global _market_map_cache, _krw_markets_cache, _market_cache_time
    now_ts = time.time()
    if (
        _market_map_cache is not None
        and _krw_markets_cache is not None
        and (now_ts - _market_cache_time) < _MARKET_CACHE_TTL
    ):
        return _market_map_cache, _krw_markets_cache
    raw = get_upbit_markets_all()
    name_map = {}
    krw_list = []
    for m in raw:
        mkt = m["market"]
        if not mkt.startswith("KRW-"):
            continue
        krw_list.append(mkt)
        korean = m.get("korean_name", "")
        english = m.get("english_name", "")
        symbol = mkt.replace("KRW-", "")
        if korean:
            name_map[korean] = mkt
        if english:
            name_map[english] = mkt
        name_map[symbol] = mkt
        name_map[mkt] = mkt
        name_map[f"{symbol}/KRW"] = mkt
    _market_map_cache = name_map
    _krw_markets_cache = krw_list
    _market_cache_time = now_ts
    return name_map, krw_list


def get_all_ticker_prices(markets):
    """전종목 시세 1회 API 호출로 조회 → { market: 현재가(int) } 반환"""
    if not markets:
        return {}
    url = "https://api.upbit.com/v1/ticker"
    try:
        resp = requests.get(url, params={"markets": ",".join(markets)}, timeout=15)
        if resp.status_code != 200:
            return {}
        data = resp.json()
        return {r["market"]: int(float(r["trade_price"])) for r in data if r.get("trade_price") is not None}
    except Exception:
        return {}


def load_excel_list(file_path):
    """upbitMA.list.xlsx 형식 엑셀 로드 (감시중=O 행만 반환)
    열: 감시중, 종목명, 감시사유, 감시가격, 감시조건, 일자, 기준가격, 비율, 수정일, 비고
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[종목별 감시] openpyxl 미설치. pip install openpyxl")
        return []
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2):
        row_dict = {}
        for idx, cell in enumerate(row):
            if idx < len(header) and header[idx]:
                row_dict[header[idx]] = cell.value
        rows.append(row_dict)
    # 감시중=O 인 행만
    active = []
    for r in rows:
        status = str(r.get("감시중", "") or "").strip().upper()
        name = str(r.get("종목명", "") or "").strip()
        if status == "O" and name:
            active.append(r)
    return active


def parse_watch_price(row):
    """행에서 감시가격 계산. 감시가격(숫자) 또는 기준가격+비율.
    반환: int 또는 None(파싱 실패/템플릿 행)
    """
    watch_raw = row.get("감시가격")
    ref_raw = row.get("기준가격")
    ratio_raw = row.get("비율")

    # 감시가격이 숫자면 사용
    if watch_raw is not None and str(watch_raw).strip() not in ("", "None", "NaT"):
        s = str(watch_raw).replace("₩", "").replace(",", "").replace("원", "").strip()
        if s and s.replace(".", "", 1).replace("-", "", 1).isdigit():
            return int(float(s))

    # 기준가격 + 비율로 계산 (기준가격이 숫자인 경우만)
    if ref_raw is None or ratio_raw is None:
        return None
    ref_str = str(ref_raw).strip()
    if not ref_str or ref_str in ("None", "NaT") or not ref_str.replace(".", "", 1).replace(",", "").replace("-", "", 1).isdigit():
        return None  # "20일선" 등 텍스트는 미지원
    try:
        ref = float(str(ref_raw).replace("₩", "").replace(",", "").replace("원", "").strip())
    except (ValueError, TypeError):
        return None
    try:
        ratio = float(str(ratio_raw).replace("%", "").strip())
    except (ValueError, TypeError):
        return None
    return int(ref * (1 + ratio / 100))


def get_current_price(market, retries=2):
    """단일 마켓 현재가 조회"""
    url = "https://api.upbit.com/v1/ticker"
    for _ in range(retries):
        try:
            resp = requests.get(url, params={"markets": market}, timeout=10)
            if resp.status_code == 200:
                data = resp.json()
                if data:
                    return int(float(data[0]["trade_price"]))
        except Exception:
            pass
        time.sleep(0.1)
    return None


def get_list_monitoring_status():
    """종목별 감시현황 메시지 본문 생성. 미사용 시 (None, 이유문자열) 반환."""
    if EXCEL_LIST_PATH is None:
        return None, "LIST_FILE 미설정"
    if not os.path.exists(EXCEL_LIST_PATH):
        return None, f"파일 없음: {EXCEL_LIST_PATH}"
    active_rows = load_excel_list(EXCEL_LIST_PATH)
    if not active_rows:
        return None, "엑셀에 감시중(O) 행 없음"
    name_market_map, _ = get_cached_market_data()
    lines = []
    count = 0
    for row in active_rows:
        stock_name = str(row.get("종목명", "") or "").strip()
        reason = str(row.get("감시사유", "") or "").strip()
        condition = str(row.get("감시조건", "") or "").strip()
        market = name_market_map.get(stock_name)
        if not market:
            for k, v in name_market_map.items():
                if k.upper() == stock_name.upper():
                    market = v
                    break
        if not market or condition not in ("이상", "이하"):
            continue
        watch_price = parse_watch_price(row)
        if watch_price is None:
            continue
        count += 1
        lines.append(f"  · {stock_name} | {reason} | {watch_price:,}원 {condition}")
    if not count:
        return "종목별 감시: 등록 0건 (엑셀 경로 있음)", None
    body = "\n".join(lines[:30])  # 최대 30건
    if count > 30:
        body += f"\n  … 외 {count - 30}건"
    return f"종목별 감시 현황 ({count}건)\n{body}", None


def run_list_monitoring():
    """LIST_FILE이 .env에 있고 해당 엑셀 파일이 있으면 종목별 감시. 전종목 시세 1회 조회 후 캐시로 비교.
    한 번 조건 충족 시 알림 전송 후 해당 (종목, 감시사유)는 감시 대상에서 제외(감시중 X와 동일)."""
    global _list_alert_sent
    if EXCEL_LIST_PATH is None or not os.path.exists(EXCEL_LIST_PATH):
        return
    active_rows = load_excel_list(EXCEL_LIST_PATH)
    if not active_rows:
        return
    name_market_map, krw_markets = get_cached_market_data()
    price_cache = get_all_ticker_prices(krw_markets)
    if not price_cache:
        print("[종목별 감시] 전종목 시세 조회 실패, 이번 주기 스킵")
        return
    now = datetime.datetime.now()
    for row in active_rows:
        stock_name = str(row.get("종목명", "") or "").strip()
        reason = str(row.get("감시사유", "") or "").strip()
        condition = str(row.get("감시조건", "") or "").strip()
        alert_key = (stock_name, reason)
        if alert_key in _list_alert_sent:
            continue

        market = name_market_map.get(stock_name)
        if not market:
            for k, v in name_market_map.items():
                if k.upper() == stock_name.upper():
                    market = v
                    break
        if not market:
            print(f"[종목별 감시] 마켓 매핑 실패: {stock_name} ({reason})")
            continue

        watch_price = parse_watch_price(row)
        if watch_price is None:
            continue  # 템플릿/비율 행 등 스킵

        if condition not in ("이상", "이하"):
            continue

        current = price_cache.get(market)
        if current is None:
            continue

        condition_met = False
        if condition == "이상":
            condition_met = current >= watch_price
        else:
            condition_met = current <= watch_price

        if not condition_met:
            continue

        _list_alert_sent.add(alert_key)

        msg = (
            f"🔔 [종목별 감시] {stock_name} - {reason}\n"
            f"   감시가격 {condition} {watch_price:,}원 | 현재가 {current:,}원\n"
            f"   ({now.strftime('%Y-%m-%d %H:%M')})"
        )
        send_telegram_message(msg)
        print(f"[종목별 감시] 알림 전송: {stock_name} ({reason})")

def get_ticker_info(markets):
    """현재가, 전일가 기준으로 등락률 계산"""
    url = "https://api.upbit.com/v1/ticker"
    res = requests.get(url, params={"markets": ",".join(markets)}).json()

    result = []
    for r in res:
        change_rate = (r['trade_price'] - r['prev_closing_price']) / r['prev_closing_price'] * 100
        result.append({
            'market': r['market'],
            'change_rate': change_rate
        })
    return result

def analyze(change_data):
    """등락률 구간별 통계 계산"""
    summary = {
        'total': len(change_data),
        'rise_5': 0,
        'rise_10': 0,
        'rise_15': 0,
        'fall_5': 0,
        'fall_10': 0,
        'fall_15': 0,
        'neutral': 0,
        'rise_over_15': [],
        'fall_below_15': []
    }

    for d in change_data:
        rate = d['change_rate']
        if rate >= 15:
            summary['rise_15'] += 1
            summary['rise_over_15'].append(d)
        if rate >= 10:
            summary['rise_10'] += 1
        if rate >= 5:
            summary['rise_5'] += 1
        if -5 < rate < 5:
            summary['neutral'] += 1
        if rate <= -5:
            summary['fall_5'] += 1
        if rate <= -10:
            summary['fall_10'] += 1
        if rate <= -15:
            summary['fall_15'] += 1
        if rate <= -15:
            summary['fall_below_15'].append(d)

    return summary

def save_to_markdown(LOGFILE, summary):
    """결과를 Markdown 파일에 추가"""
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    lines = []
    lines.append(f"\n# 📈 업비트 원화시장 상승/하락 통계 ({now})\n")
    lines.append("| 구분 | 종목 수 |")
    lines.append("|------|----------|")
    lines.append(f"| 전체 종목 | {summary['total']} |")
    lines.append(f"| (+15% 이상) | {summary['rise_15']} |")
    lines.append(f"| (+10% 이상) | {summary['rise_10']} |")
    lines.append(f"| +5% 이상 | {summary['rise_5']} |")
    lines.append(f"| -5% ~ +5% | {summary['neutral']} |")
    lines.append(f"| -5% 이하 | {summary['fall_5']} |")
    lines.append(f"| (-10% 이하) | {summary['fall_10']} |")
    lines.append(f"| (-15% 이하) | {summary['fall_15']} |")

    lines.append("\n## 🚀 +15% 이상 상승 종목")
    if summary['rise_over_15']:
        lines.append("| 종목명 | 상승률(%) |")
        lines.append("|--------|------------|")
        for d in summary['rise_over_15']:
            lines.append(f"| {d['market']} | {d['change_rate']:.2f}% |")
    else:
        lines.append("- 없음")

    lines.append("\n## 📉 -15% 이하 하락 종목")
    if summary['fall_below_15']:
        lines.append("| 종목명 | 하락률(%) |")
        lines.append("|--------|------------|")
        for d in summary['fall_below_15']:
            lines.append(f"| {d['market']} | {d['change_rate']:.2f}% |")
    else:
        lines.append("- 없음")

    with open(LOGFILE, "a", encoding="utf-8") as f:
        f.write("\n".join(lines))
        f.write("\n\n---\n\n")

    print(f"[{now}] Markdown 파일 저장 완료 → {LOGFILE}")
    return len(summary['fall_below_15'])

def main():
    now_start = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    send_telegram_message(
        f"🟢 [upbitMA] 업비트 원화시장 감시 스크립트 시작\n({now_start})"
    )
    print(f"[시작] 텔레그램 알림 전송 완료 → {now_start}")

    def on_exit():
        t = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        send_telegram_message(f"🔴 [upbitMA] 스크립트 종료\n({t})")

    atexit.register(on_exit)
    signal.signal(signal.SIGINT, lambda s, f: (on_exit(), sys.exit(0)))
    signal.signal(signal.SIGTERM, lambda s, f: (on_exit(), sys.exit(0)))

    last_daily_report_date = None  # 매일 8:30 리포트 중복 방지
    last_full_analysis_time = None  # 전체 종목 분석 마지막 실행 시각

    while True:
        try:
            now = datetime.datetime.now()
            hour, minute = now.hour, now.minute
            today = now.date()

            # 전체 종목 분석은 ALL_MA_INTERVAL(기본 1시간)마다만 실행
            do_full_analysis = (
                last_full_analysis_time is None
                or (now - last_full_analysis_time).total_seconds() >= ALL_MA_INTERVAL
            )

            if do_full_analysis:
                # === 데이터 수집 및 분석 (1시간 단위) ===
                markets = get_upbit_markets()
                change_data = get_ticker_info(markets)
                summary = analyze(change_data)
                fall_count = save_to_markdown(LOG_DIR_FILENAME, summary)
                last_full_analysis_time = now

                # === ① 이벤트: -15% 이하 하락 15개 이상 시에만 텔레그램 전송 ===
                if fall_count >= 15:
                    msg = (
                        f"📉 경고: -15% 이하 하락 종목이 {fall_count}개 이상 발생!\n"
                        f"({now.strftime('%Y-%m-%d %H:%M')})\n"
                        f"전체 종목: {summary['total']}개\n"
                        f"상승: +5%↑ {summary['rise_5']}개 (+10%↑ {summary['rise_10']}개 | +15%↑ {summary['rise_15']}개)\n"
                        f"보합(-5%~+5%): {summary['neutral']}개\n"
                        f"하락: -5%↓ {summary['fall_5']}개 (-10%↓ {summary['fall_10']}개 | -15%↓ {summary['fall_15']}개)\n"
                        f"파일: {os.path.basename(LOG_DIR_FILENAME)}"
                    )
                    send_telegram_message(msg)

                # === ② 매일 8:30 정리 리포트 (해당일 1회만 텔레그램 전송) ===
                is_after_830 = (hour > 8) or (hour == 8 and minute >= 30)
                if is_after_830 and last_daily_report_date != today:
                    msg_summary = (
                        f"📊 업비트 원화시장 요약 리포트 ({now.strftime('%Y-%m-%d %H:%M')})\n"
                        f"전체 종목: {summary['total']}개\n"
                        f"상승: +5%↑ {summary['rise_5']}개 (+10%↑ {summary['rise_10']}개 | +15%↑ {summary['rise_15']}개)\n"
                        f"보합(-5%~+5%): {summary['neutral']}개\n"
                        f"하락: -5%↓ {summary['fall_5']}개 (-10%↓ {summary['fall_10']}개 | -15%↓ {summary['fall_15']}개)\n"
                        f"파일: {os.path.basename(LOG_DIR_FILENAME)}"
                    )
                    send_telegram_message(msg_summary)
                    last_daily_report_date = today
                    print(f"[로그] 매일 8:30 정리 리포트 전송 완료 ({now.strftime('%Y-%m-%d %H:%M')})")

                # === 종목별 감시현황: 로그에만 (1시간마다) ===
                try:
                    status, reason = get_list_monitoring_status()
                    if status:
                        print(f"[로그] 종목별 감시 현황: {status[:80]}..." if len(status) > 80 else f"[로그] 종목별 감시 현황: {status}")
                    else:
                        print(f"[로그] 종목별 감시: {reason}")
                except Exception as e_status:
                    print(f"[종목별 감시현황 오류] {e_status}")

            # === ③ 종목별 주가 감시 (1분 단위, 감시가 도달 시에만 텔레그램) ===
            try:
                run_list_monitoring()
            except Exception as e_list:
                print(f"[종목별 감시 오류] {e_list}")

        except Exception as e:
            print(f"[오류 발생] {e}")

        print(f"⏳ {LIST_MA_INTERVAL}초 대기 중...\n")
        time.sleep(LIST_MA_INTERVAL)

if __name__ == "__main__":
    main()
