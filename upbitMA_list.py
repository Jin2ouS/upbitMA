# upbitMA_list.py - 리스트(종목별) 감시 전용 (upbitMA.list.xlsx 기반)
# created : 2026-02-03 (upbitMA 분리)
# 수정: .env LIST_FILE, LIST_MA_INTERVAL 사용

import os
import sys
import time
import datetime
import atexit
import signal

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

from dotenv import load_dotenv

from utils_upbit import send_telegram_message, get_upbit_markets_all, get_all_ticker_prices

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(SCRIPT_DIR, ".env"))

LIST_MA_INTERVAL = int(os.getenv("LIST_MA_INTERVAL", "60").strip() or "60")
LIST_FILE_RAW = os.getenv("LIST_FILE", "").strip()
if LIST_FILE_RAW:
    EXCEL_LIST_PATH = os.path.join(SCRIPT_DIR, LIST_FILE_RAW) if not os.path.isabs(LIST_FILE_RAW) else LIST_FILE_RAW
else:
    EXCEL_LIST_PATH = None

if not os.getenv("TELEGRAM_BOT_TOKEN", "").strip() or not os.getenv("TELEGRAM_CHAT_ID", "").strip():
    raise ValueError("TELEGRAM_BOT_TOKEN, TELEGRAM_CHAT_ID가 .env에 필요합니다.")

# 리스트 감시용 캐시
_MARKET_CACHE_TTL = 600
_market_map_cache = None
_krw_markets_cache = None
_market_cache_time = 0

_list_alert_sent = set()
_last_active_list_count = 0


def get_cached_market_data():
    """종목명 매핑 + KRW 마켓 목록 캐시. TTL 내에는 API 호출 없이 반환."""
    global _market_map_cache, _krw_markets_cache, _market_cache_time
    now_ts = time.time()
    if (
        _market_map_cache is not None
        and _krw_markets_cache is not None
        and (now_ts - _market_cache_time) < _MARKET_CACHE_TTL
    ):
        return _market_map_cache, _krw_markets_cache
    raw = get_upbit_markets_all()
    name_map = {}
    krw_list = []
    for m in raw:
        mkt = m["market"]
        if not mkt.startswith("KRW-"):
            continue
        krw_list.append(mkt)
        korean = m.get("korean_name", "")
        english = m.get("english_name", "")
        symbol = mkt.replace("KRW-", "")
        if korean:
            name_map[korean] = mkt
        if english:
            name_map[english] = mkt
        name_map[symbol] = mkt
        name_map[mkt] = mkt
        name_map[f"{symbol}/KRW"] = mkt
    _market_map_cache = name_map
    _krw_markets_cache = krw_list
    _market_cache_time = now_ts
    return name_map, krw_list


def load_excel_list(file_path):
    """upbitMA.list.xlsx 형식 엑셀 로드 (감시중=O 행만 반환)"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[리스트 감시] openpyxl 미설치. pip install openpyxl")
        return []
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2):
        row_dict = {}
        for idx, cell in enumerate(row):
            if idx < len(header) and header[idx]:
                row_dict[header[idx]] = cell.value
        rows.append(row_dict)
    active = []
    for r in rows:
        status = str(r.get("감시중", "") or "").strip().upper()
        name = str(r.get("종목명", "") or "").strip()
        if status == "O" and name:
            active.append(r)
    return active


def parse_list_price(row):
    """행에서 감시가격 계산. 감시가격(숫자) 또는 기준가격+비율."""
    list_price_raw = row.get("감시가격")
    ref_raw = row.get("기준가격")
    ratio_raw = row.get("비율")

    if list_price_raw is not None and str(list_price_raw).strip() not in ("", "None", "NaT"):
        s = str(list_price_raw).replace("₩", "").replace(",", "").replace("원", "").strip()
        if s and s.replace(".", "", 1).replace("-", "", 1).isdigit():
            return int(float(s))

    if ref_raw is None or ratio_raw is None:
        return None
    ref_str = str(ref_raw).strip()
    if (
        not ref_str
        or ref_str in ("None", "NaT")
        or not ref_str.replace(".", "", 1).replace(",", "").replace("-", "", 1).isdigit()
    ):
        return None
    try:
        ref = float(str(ref_raw).replace("₩", "").replace(",", "").replace("원", "").strip())
    except (ValueError, TypeError):
        return None
    try:
        ratio = float(str(ratio_raw).replace("%", "").strip())
    except (ValueError, TypeError):
        return None
    return int(ref * (1 + ratio / 100))


def get_list_monitoring_status():
    """리스트 감시 현황 메시지 본문 생성."""
    if EXCEL_LIST_PATH is None:
        return None, "LIST_FILE 미설정"
    if not os.path.exists(EXCEL_LIST_PATH):
        return None, f"파일 없음: {EXCEL_LIST_PATH}"
    active_rows = load_excel_list(EXCEL_LIST_PATH)
    if not active_rows:
        return None, "엑셀에 감시중(O) 행 없음"
    name_market_map, _ = get_cached_market_data()
    lines = []
    count = 0
    for row in active_rows:
        stock_name = str(row.get("종목명", "") or "").strip()
        reason = str(row.get("감시사유", "") or "").strip()
        condition = str(row.get("감시조건", "") or "").strip()
        market = name_market_map.get(stock_name)
        if not market:
            for k, v in name_market_map.items():
                if k.upper() == stock_name.upper():
                    market = v
                    break
        if not market or condition not in ("이상", "이하"):
            continue
        list_price = parse_list_price(row)
        if list_price is None:
            continue
        count += 1
        lines.append(f"  · {stock_name} | {reason} | {list_price:,}원 {condition}")
    if not count:
        return "리스트 감시: 등록 0건 (엑셀 경로 있음)", None
    body = "\n".join(lines[:30])
    if count > 30:
        body += f"\n  … 외 {count - 30}건"
    return f"리스트 감시 현황 ({count}건)\n{body}", None


def run_list_monitoring():
    """리스트 감시 실행. 조건 충족 시 알림 후 해당 (종목, 감시사유)는 감시 대상에서 제외."""
    global _list_alert_sent, _last_active_list_count
    if EXCEL_LIST_PATH is None or not os.path.exists(EXCEL_LIST_PATH):
        return
    active_rows = load_excel_list(EXCEL_LIST_PATH)
    if not active_rows:
        return
    _last_active_list_count = len(active_rows)
    name_market_map, krw_markets = get_cached_market_data()
    price_cache = get_all_ticker_prices(krw_markets)
    if not price_cache:
        print("[리스트 감시] 전종목 시세 조회 실패, 이번 주기 스킵")
        return
    now = datetime.datetime.now()
    for row in active_rows:
        stock_name = str(row.get("종목명", "") or "").strip()
        reason = str(row.get("감시사유", "") or "").strip()
        condition = str(row.get("감시조건", "") or "").strip()
        alert_key = (stock_name, reason)
        if alert_key in _list_alert_sent:
            continue

        market = name_market_map.get(stock_name)
        if not market:
            for k, v in name_market_map.items():
                if k.upper() == stock_name.upper():
                    market = v
                    break
        if not market:
            print(f"[리스트 감시] 마켓 매핑 실패: {stock_name} ({reason})")
            continue

        list_price = parse_list_price(row)
        if list_price is None:
            continue
        if condition not in ("이상", "이하"):
            continue

        current = price_cache.get(market)
        if current is None:
            continue

        condition_met = False
        if condition == "이상":
            condition_met = current >= list_price
        else:
            condition_met = current <= list_price

        if not condition_met:
            continue

        _list_alert_sent.add(alert_key)
        msg = (
            f"🔔 [리스트 감시] {stock_name} - {reason}\n"
            f"   감시가격 {condition} {list_price:,}원 | 현재가 {current:,}원\n"
            f"   ({now.strftime('%Y-%m-%d %H:%M')})"
        )
        send_telegram_message(msg)
        print(f"[리스트 감시] 알림 전송: {stock_name} ({reason})")


def main():
    now_start = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    send_telegram_message(f"🟢 [upbitMA_list] 리스트 감시 스크립트 시작\n({now_start})")
    print(f"[시작] 텔레그램 알림 전송 완료 → {now_start}")

    def on_exit():
        t = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        send_telegram_message(f"🔴 [upbitMA_list] 스크립트 종료\n({t})")

    atexit.register(on_exit)
    signal.signal(signal.SIGINT, lambda s, f: (on_exit(), sys.exit(0)))
    signal.signal(signal.SIGTERM, lambda s, f: (on_exit(), sys.exit(0)))

    first_list_status_telegram_sent = False

    while True:
        try:
            # 최초 1회: 리스트 감시 현황 텔레그램 전송
            if not first_list_status_telegram_sent:
                try:
                    status, reason = get_list_monitoring_status()
                    if status:
                        send_telegram_message(f"📋 [upbitMA_list] {status}")
                    else:
                        send_telegram_message(f"📋 [upbitMA_list] 리스트 감시: 미사용 ({reason})")
                    first_list_status_telegram_sent = True
                except Exception as e_status:
                    print(f"[리스트 감시 현황 오류] {e_status}")

            run_list_monitoring()
        except Exception as e:
            print(f"[오류 발생] {e}")

        now = datetime.datetime.now()
        next_run = now + datetime.timedelta(seconds=LIST_MA_INTERVAL)
        list_active_count = max(0, _last_active_list_count - len(_list_alert_sent))
        excluded = len(_list_alert_sent)
        print(
            f"[{now.strftime('%H:%M:%S')}] ⏳ {LIST_MA_INTERVAL}초 대기 중... "
            f"다음 {next_run.strftime('%H:%M:%S')} | 리스트 {list_active_count}건 | 제외 {excluded}건"
        )
        time.sleep(LIST_MA_INTERVAL)


if __name__ == "__main__":
    main()
